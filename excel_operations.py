import os

from openpyxl import load_workbook

from parse_base import find_base


def write_status_in_excel(code: int, folder_path: str, counter: int):
    """Запись статуса выполнения загрузки в базу"""

    if code == 200:
        base_path = find_base(folder_path)
        wb = load_workbook(base_path)
        sheet = wb[wb.active.title]
        sheet.cell(row=1, column=6, value="Статус выполнения")
        sheet.cell(row=counter, column=6, value="Загружено в Хантфлоу")
        os.chdir(folder_path)
        wb.save("Тестовая база.xlsx")


def check_status_in_excel(folder_path: str):
    """Проверка статуса выполнения загрузки и возращение области Excel по которой необходимо загрузить кандидатов"""

    base_path = find_base(folder_path)
    wb = load_workbook(base_path)
    sheet = wb[wb.active.title]
    min_row = sheet.min_row
    min_column = sheet.min_column
    min_coordinate = sheet.cell(row=min_row, column=min_column).coordinate
    max_row = sheet.max_row
    max_column = sheet.max_column
    max_coordinate = sheet.cell(row=max_row, column=max_column).coordinate
    if sheet.cell(row=max_row, column=max_column).value == "Загружено в Хантфлоу":
        return "База уже загружена в Хантфлоу"
    find_zone = []
    for cellObj in sheet[min_coordinate:max_coordinate]:
        for cell in cellObj:
            if cell.value == "Статус выполнения":
                column = cell.column
                row = cell.row
                for i in range(row+1, max_row+1):
                    if sheet.cell(row=i, column=column).value != "Загружено в Хантфлоу":
                        find_zone.append(sheet.cell(row=i, column=column-column+1).coordinate)
                        break
    if len(find_zone) != 0:
        response = {}
        response["min_coordinate"] = find_zone[0]
        response["max_coordinate"] = max_coordinate
        response["counter"] = int(find_zone[0][1])
        return response
    else:
        response = {}
        min_coordinate = sheet.cell(row=min_row+1, column=min_column).coordinate
        response["min_coordinate"] = min_coordinate
        response["max_coordinate"] = max_coordinate
        response["counter"] = min_row+1
        return response
